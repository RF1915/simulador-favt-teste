import streamlit as st
import pandas as pd
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import io
import datetime
from typing import Dict, Any, Optional

XLS_PATH = "FAVT-V41.xlsx"


# ----------------------------
# Helpers
# ----------------------------
def _norm(x):
    if x is None:
        return ""
    return str(x).strip()


def _sheet_df(ws, start_row: int, start_col: int, end_col: int, end_row: int = None):
    rows = []
    r = start_row
    while True:
        if end_row is not None and r > end_row:
            break
        vals = [ws.cell(row=r, column=c).value for c in range(start_col, end_col + 1)]
        if end_row is None and all(v is None for v in vals):
            break
        rows.append(vals)
        r += 1
        if end_row is None and r > start_row + 5000:
            break
    return pd.DataFrame(rows)


def vlookup_exact(key, df: pd.DataFrame, key_col_idx: int, return_col_idx: int):
    for _, row in df.iterrows():
        if row.iloc[key_col_idx] == key:
            return row.iloc[return_col_idx]
    raise KeyError(f"VLOOKUP key not found: {key}")


# ----------------------------
# Load Excel tables
# ----------------------------
def load_tables(xls_path: str):
    wb = openpyxl.load_workbook(xls_path, data_only=False)

    # PREÇO BAIRROS: headers row4 A:G, data A5:G92
    ws = wb["PREÇO BAIRROS"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, 8)]
    rows = []
    for r in range(5, 93):
        row = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        if all(v is None for v in row):
            continue
        rows.append(row)
    preco_df = pd.DataFrame(rows, columns=[_norm(h) for h in headers])

    # Factor tables
    f1_df = _sheet_df(wb["Factor1_H"], start_row=11, start_col=1, end_col=5)
    f2_df = _sheet_df(wb["Factor2_Elev"], start_row=11, start_col=1, end_col=4)
    f3_hist_df = _sheet_df(wb["Factor3_proporção"], start_row=30, start_col=5, end_col=7)    # E:G
    f3_nhist_df = _sheet_df(wb["Factor3_proporção"], start_row=30, start_col=11, end_col=13)  # K:M
    f4_df = _sheet_df(wb["Factor4_Uso"], start_row=5, start_col=5, end_col=24)                 # E:X
    f4_map_df = _sheet_df(wb["Factor4_Uso"], start_row=6, start_col=28, end_col=30)            # AB:AD
    vistas_df = _sheet_df(wb["N_Vistas"], start_row=3, start_col=2, end_col=3)                  # B:C
    infra_df = _sheet_df(wb["Factor5_infra"], start_row=8, start_col=2, end_col=3)              # B:C

    return {
        "preco_bairros": preco_df,
        "factor1": f1_df,
        "factor2": f2_df,
        "factor3_hist": f3_hist_df,
        "factor3_nhist": f3_nhist_df,
        "factor4": f4_df,
        "factor4_map": f4_map_df,
        "vistas": vistas_df,
        "infra": infra_df,
    }


def build_reference_maps(preco_df: pd.DataFrame):
    """
    Objetivo:
      - "Zona" (territorial) = coluna para agrupar/filtrar bairros (UX)
      - "Bairro" = nome do bairro/localidade
      - "Ponderação" = Alta/Média/Baixa = colunas numéricas de preço (cálculo)
    """
    cols = list(preco_df.columns)

    # Estrutura típica do ficheiro:
    # A = código
    # B = Zona (territorial)  -> pode chamar "Zona", "ZONA", etc.
    # C = Bairro/Localidade
    code_col = cols[0]
    bairro_col = cols[2] if len(cols) >= 3 else cols[-1]

    # tentar detetar "Zona" territorial por nome; fallback: coluna B
    zona_col: Optional[str] = None
    for c in cols:
        name = str(c).strip().lower()
        if ("zona" in name) and ("hist" not in name) and ("ponder" not in name) and ("valor" not in name):
            # evita confundir com "Zona Histórica" ou algo semelhante
            if c != bairro_col and c != code_col:
                zona_col = c
                break
    if zona_col is None and len(cols) >= 2:
        zona_col = cols[1]

    # coluna histórica (se existir)
    hist_col = None
    for c in cols:
        if "hist" in str(c).lower():
            hist_col = c
            break

    # Ponderação/Valorização = colunas de preço (Alta/Média/Baixa)
    ponder_whitelist = {"alta", "média", "media", "baixa", "alta(a)", "média(m)", "media(m)", "baixa(b)"}
    ponder_cols = []
    for c in cols:
        if c in {code_col, bairro_col, zona_col, hist_col}:
            continue
        if str(c).strip().lower() in ponder_whitelist:
            ponder_cols.append(c)

    # fallback: detetar colunas maioritariamente numéricas
    if not ponder_cols:
        for c in cols:
            if c in {code_col, bairro_col, zona_col, hist_col}:
                continue
            s = pd.to_numeric(preco_df[c], errors="coerce")
            if s.notna().mean() > 0.8:
                ponder_cols.append(c)

    # Mapas
    bairro_to_code = dict(zip(preco_df[bairro_col].astype(str), preco_df[code_col]))
    bairro_to_hist = {}
    if hist_col is not None:
        bairro_to_hist = dict(zip(preco_df[bairro_col].astype(str), preco_df[hist_col]))

    bairro_to_zona = {}
    if zona_col is not None:
        bairro_to_zona = dict(zip(preco_df[bairro_col].astype(str), preco_df[zona_col]))

    return {
        "bairro_col": bairro_col,
        "code_col": code_col,
        "zona_col": zona_col,
        "hist_col": hist_col,
        "ponder_cols": ponder_cols,           # colunas de preço para o cálculo
        "bairro_to_code": bairro_to_code,
        "bairro_to_hist": bairro_to_hist,
        "bairro_to_zona": bairro_to_zona,     # agrupamento territorial para UX
    }


# ----------------------------
# Cálculo FAVT
# ----------------------------
def calc_favt(payload: Dict[str, Any], tables: Dict[str, pd.DataFrame], ref: Dict[str, Any], mode: str):
    preco_df = tables["preco_bairros"]

    zona_territorial = payload["zona_territorial"]
    bairro = payload["bairro"]
    ponderacao = payload["ponderacao"]  # Alta/Média/Baixa (coluna de preço)

    area_total = float(payload["area_total"])
    area_coberta = float(payload["area_coberta"])
    n_vistas = int(payload["n_vistas"])
    n_pisos_in = int(payload["n_pisos"])
    largura_infra = float(payload["largura_infra"])
    largura_fachada = float(payload["largura_fachada"])
    uso = payload["uso"]

    # regra especial do FAVT2 (como no XLS que descreveste)
    if mode == "FAVT2":
        n_pisos_favt1 = int(payload.get("n_pisos_favt1_for_rule", n_pisos_in))
        n_pisos = 3 if (n_pisos_favt1 == 1 and n_pisos_in == 2) else n_pisos_in
    else:
        n_pisos = n_pisos_in

    bairro_code = ref["bairro_to_code"].get(str(bairro))
    if bairro_code is None:
        raise ValueError(f"Bairro não encontrado na tabela: {bairro}")

    # validação adicional: bairro pertence à zona escolhida (se houver mapeamento)
    if ref.get("bairro_to_zona"):
        z = ref["bairro_to_zona"].get(str(bairro))
        if (z is not None) and (str(z).strip() != str(zona_territorial).strip()):
            # não impede (porque pode haver diferenças no Excel), mas avisa com erro para manter consistência
            raise ValueError(f"Bairro '{bairro}' não pertence à Zona '{zona_territorial}' (segundo a tabela).")

    # Zona histórica
    zona_historica = payload.get("zona_historica")
    if not zona_historica:
        zona_historica = ref["bairro_to_hist"].get(str(bairro), "Não")
    zona_historica = "Sim" if str(zona_historica).strip().lower().startswith("s") else "Não"

    indice_ocupacao = (area_coberta / area_total) if area_total else 0.0
    proporcao = (largura_fachada / n_pisos) if n_pisos else 0.0

    # preço base (ponderação = coluna de preço)
    if ponderacao not in preco_df.columns:
        raise ValueError(f"Ponderação '{ponderacao}' inválida. Disponíveis: {ref['ponder_cols']}")
    match = preco_df[preco_df[ref["code_col"]] == bairro_code]
    if match.empty:
        raise ValueError(f"Código '{bairro_code}' não encontrado em PREÇO BAIRROS.")
    preco_base = float(match.iloc[0][ponderacao])

    # F1
    f1 = float(vlookup_exact(n_pisos, tables["factor1"], 0, 4)) if n_pisos > 1 else 0.0
    # F2
    f2 = float(vlookup_exact(n_pisos, tables["factor2"], 0, 3)) if n_pisos > 1 else 0.0

    # F3
    if (zona_historica == "Sim") and (proporcao < 5):
        f3 = float(vlookup_exact(n_pisos, tables["factor3_hist"], 0, 2))
    else:
        add_bool = 1.0 if ((zona_historica == "Não") and (proporcao < 2)) else 0.0
        f3 = add_bool + float(vlookup_exact(n_pisos, tables["factor3_nhist"], 0, 2))

    # F4 (uso)
    f4_map = tables["factor4_map"]
    try:
        uso_code = vlookup_exact(uso, f4_map, 0, 2)
    except KeyError:
        uso_code = None
        for _, row in f4_map.iterrows():
            if str(row.iloc[0]).strip().lower() in str(uso).strip().lower():
                uso_code = row.iloc[2]
                break
        if uso_code is None:
            raise ValueError(f"Uso não encontrado no mapeamento: {uso}")

    f4_table = tables["factor4"]
    headers = list(f4_table.iloc[0].values)
    if uso_code not in headers:
        raise ValueError(f"Código de uso '{uso_code}' não existe nos cabeçalhos Factor4_Uso.")
    col_idx = headers.index(uso_code)
    body = f4_table.iloc[1:].reset_index(drop=True)
    f4 = float(vlookup_exact(uso_code, body, 0, col_idx))

    # F5
    f5 = 0.0 if largura_infra < 10 else float(vlookup_exact(largura_infra, tables["infra"], 0, 1))

    total_ponderacao = (1 + f1) * (1 + f2) * (1 + f3) * (1 + f4) * (1 + f5)
    preco_ponderado = total_ponderacao * preco_base

    # Ajuste área
    ajuste_area = 0.0
    if area_total <= 200:
        ajuste_area = 0.0
    elif 200 < area_total < 1000:
        cand = (indice_ocupacao - 0.8) * preco_base
        ajuste_area = cand if cand > 0 else 0.0
    elif area_total > 1000:
        cand = (indice_ocupacao - 0.4) * preco_base
        ajuste_area = cand if cand > 0 else 0.0

    fator_vistas = float(vlookup_exact(n_vistas, tables["vistas"], 0, 1))

    preco_sem_vistas = preco_ponderado + ajuste_area
    preco_com_vistas = (1 + fator_vistas) * preco_sem_vistas
    valor_terreno = preco_com_vistas * area_total

    return {
        "preco_base": preco_base,
        "fatores": {"f1": f1, "f2": f2, "f3": f3, "f4": f4, "f5": f5},
        "total_ponderacao": total_ponderacao,
        "indice_ocupacao": indice_ocupacao,
        "ajuste_area": ajuste_area,
        "fator_vistas": fator_vistas,
        "preco_total_sem_vistas": preco_sem_vistas,
        "preco_total_com_vistas": preco_com_vistas,
        "valor_terreno": valor_terreno,
        "inputs_norm": {
            "zona_territorial": zona_territorial,
            "bairro": bairro,
            "ponderacao": ponderacao,
            "area_total": area_total,
            "area_coberta": area_coberta,
            "n_vistas": n_vistas,
            "n_pisos": n_pisos,
            "largura_infra": largura_infra,
            "largura_fachada": largura_fachada,
            "uso": uso,
            "zona_historica": zona_historica,
        },
    }


# ----------------------------
# PDF
# ----------------------------
def render_pdf(result: Dict[str, Any], titulo: str):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    y = h - 20 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, y, titulo)
    y -= 8 * mm

    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, y, f"Data/Hora: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 10 * mm

    inp = result["inputs_norm"]
    lines = [
        f"Zona (territorial): {inp['zona_territorial']}",
        f"Bairro/Localidade: {inp['bairro']}",
        f"Ponderação/Valorização: {inp['ponderacao']}",
        f"Área Total (m²): {inp['area_total']:.2f}",
        f"Área Coberta (m²): {inp['area_coberta']:.2f}",
        f"Nº de Pisos: {inp['n_pisos']}",
        f"Nº de Vistas: {inp['n_vistas']}",
        f"Largura do lote / Infra (m): {inp['largura_infra']:.2f}",
        f"Largura da fachada (m): {inp['largura_fachada']:.2f}",
        f"Uso: {inp['uso']}",
        f"Zona histórica: {inp['zona_historica']}",
    ]
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y, "Entradas")
    y -= 6 * mm
    c.setFont("Helvetica", 10)
    for ln in lines:
        c.drawString(22 * mm, y, ln)
        y -= 5 * mm

    y -= 4 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y, "Resultados")
    y -= 6 * mm
    c.setFont("Helvetica", 10)

    f = result["fatores"]
    out_lines = [
        f"Preço base ($/m²): {result['preco_base']:.2f}",
        f"F1: {f['f1']:.4f}   F2: {f['f2']:.4f}   F3: {f['f3']:.4f}   F4: {f['f4']:.4f}   F5: {f['f5']:.4f}",
        f"Total ponderação: {result['total_ponderacao']:.6f}",
        f"Índice de ocupação: {result['indice_ocupacao']:.6f}",
        f"Ajuste área ($/m²): {result['ajuste_area']:.2f}",
        f"Fator vistas: {result['fator_vistas']:.4f}",
        f"Preço total ponderado ($/m²): {result['preco_total_com_vistas']:.2f}",
        f"Valor do terreno ($): {result['valor_terreno']:.2f}",
    ]
    for ln in out_lines:
        c.drawString(22 * mm, y, ln)
        y -= 5 * mm

    c.showPage()
    c.save()
    buf.seek(0)
    return buf.getvalue()


# ----------------------------
# App UI
# ----------------------------
st.set_page_config(page_title="Simulador FAVT (teste)", layout="wide")
st.title("Simulador FAVT (site de teste)")
st.caption("Fluxo igual ao site CMP: Zona → Bairro/Localidade → Ponderação (Alta/Média/Baixa).")

@st.cache_resource
def _load():
    tables = load_tables(XLS_PATH)
    ref = build_reference_maps(tables["preco_bairros"])
    return tables, ref

tables, ref = _load()

# Zonas territoriais (para UX). Se não existir, cria uma única "—"
if ref.get("bairro_to_zona"):
    zonas_territoriais = sorted({str(z) for z in ref["bairro_to_zona"].values() if str(z).strip() != ""})
else:
    zonas_territoriais = ["—"]

# Ponderações do cálculo (colunas de preço)
# Mantemos “Alta/Média/Baixa” por padrão e tentamos casar com os nomes reais no Excel.
def _pick_ponder_label(label: str) -> str:
    # tenta encontrar a coluna exata no Excel
    for c in ref["ponder_cols"]:
        if str(c).strip().lower() == label.lower():
            return c
    # tenta versões com acento / sem acento
    if label.lower() == "média":
        for c in ref["ponder_cols"]:
            if str(c).strip().lower() in {"media", "média"}:
                return c
    return label

ponder_ui = ["Alta", "Média", "Baixa"]

usos = [u for u in tables["factor4_map"].iloc[:, 0].dropna().astype(str).tolist()]

col1, col2 = st.columns(2)

with col1:
    st.subheader("FAVT 1 (Original)")

    zt1 = st.selectbox("Zona (FAVT1)", zonas_territoriais, key="zt1")
    # filtra bairros por zona
    bairros_all = sorted(list(ref["bairro_to_code"].keys()))
    if zt1 != "—" and ref.get("bairro_to_zona"):
        bairros_1 = [b for b in bairros_all if str(ref["bairro_to_zona"].get(b, "")).strip() == str(zt1).strip()]
        if not bairros_1:
            bairros_1 = bairros_all
    else:
        bairros_1 = bairros_all

    bairro1 = st.selectbox("Bairro/Localidade (FAVT1)", bairros_1, key="b1")

    pond1_ui = st.selectbox("Ponderação/Valorização (FAVT1)", ponder_ui, key="pnd1")
    pond1 = _pick_ponder_label(pond1_ui)

    area_total1 = st.number_input("Área total A (m²)", min_value=1.0, value=250.0, step=1.0, key="a1")
    area_coberta1 = st.number_input("Área coberta B (m²)", min_value=0.0, value=220.0, step=1.0, key="b1a")
    n_vistas1 = st.selectbox("Nº de vistas", [1, 2, 3, 4], index=0, key="v1")
    n_pisos1 = st.number_input("Nº de pisos", min_value=1, value=2, step=1, key="pis1")
    largura_infra1 = st.number_input("Largura do lote / Infra (m)", min_value=0.0, value=25.0, step=0.5, key="i1")
    largura_fachada1 = st.number_input("Largura da fachada (m)", min_value=0.0, value=10.0, step=0.5, key="f1")
    uso1 = st.selectbox("Uso", usos, key="u1")

    zona_hist1_default = "Sim" if str(ref["bairro_to_hist"].get(bairro1, "Não")).strip().lower().startswith("s") else "Não"
    zona_hist1 = st.selectbox("Zona Histórica? (override)", ["(auto)", "Sim", "Não"], index=0, key="h1")
    zona_hist1_val = zona_hist1_default if zona_hist1 == "(auto)" else zona_hist1

with col2:
    st.subheader("FAVT 2 (Alterado)")

    zt2 = st.selectbox("Zona (FAVT2)", zonas_territoriais, key="zt2",
                       index=zonas_territoriais.index(zt1) if zt1 in zonas_territoriais else 0)

    if zt2 != "—" and ref.get("bairro_to_zona"):
        bairros_2 = [b for b in bairros_all if str(ref["bairro_to_zona"].get(b, "")).strip() == str(zt2).strip()]
        if not bairros_2:
            bairros_2 = bairros_all
    else:
        bairros_2 = bairros_all

    bairro2 = st.selectbox("Bairro/Localidade (FAVT2)", bairros_2, key="b2",
                           index=bairros_2.index(bairro1) if bairro1 in bairros_2 else 0)

    pond2_ui = st.selectbox("Ponderação/Valorização (FAVT2)", ponder_ui, key="pnd2",
                            index=ponder_ui.index(pond1_ui) if pond1_ui in ponder_ui else 0)
    pond2 = _pick_ponder_label(pond2_ui)

    area_total2 = st.number_input("Área total A (m²) (FAVT2)", min_value=1.0, value=float(area_total1), step=1.0, key="a2")
    area_coberta2 = st.number_input("Área coberta B (m²) (FAVT2)", min_value=0.0, value=float(area_coberta1), step=1.0, key="b2a")
    n_vistas2 = st.selectbox("Nº de vistas (FAVT2)", [1, 2, 3, 4], index=[1, 2, 3, 4].index(n_vistas1), key="v2")
    n_pisos2 = st.number_input("Nº de pisos (FAVT2)", min_value=1, value=4, step=1, key="pis2")
    largura_infra2 = st.number_input("Largura do lote / Infra (m) (FAVT2)", min_value=0.0, value=float(largura_infra1), step=0.5, key="i2")
    largura_fachada2 = st.number_input("Largura da fachada (m) (FAVT2)", min_value=0.0, value=float(largura_fachada1), step=0.5, key="f2")
    uso2 = st.selectbox("Uso (FAVT2)", usos, key="u2", index=usos.index(uso1) if uso1 in usos else 0)

    zona_hist2_default = "Sim" if str(ref["bairro_to_hist"].get(bairro2, "Não")).strip().lower().startswith("s") else "Não"
    zona_hist2 = st.selectbox("Zona Histórica? (override) (FAVT2)", ["(auto)", "Sim", "Não"], index=0, key="h2")
    zona_hist2_val = zona_hist2_default if zona_hist2 == "(auto)" else zona_hist2

st.divider()
run = st.button("Simular")

if run:
    try:
        res1 = calc_favt({
            "zona_territorial": zt1,
            "bairro": bairro1,
            "ponderacao": pond1,
            "area_total": area_total1,
            "area_coberta": area_coberta1,
            "n_vistas": n_vistas1,
            "n_pisos": n_pisos1,
            "largura_infra": largura_infra1,
            "largura_fachada": largura_fachada1,
            "uso": uso1,
            "zona_historica": zona_hist1_val,
        }, tables, ref, "FAVT1")

        res2 = calc_favt({
            "zona_territorial": zt2,
            "bairro": bairro2,
            "ponderacao": pond2,
            "area_total": area_total2,
            "area_coberta": area_coberta2,
            "n_vistas": n_vistas2,
            "n_pisos": n_pisos2,
            "largura_infra": largura_infra2,
            "largura_fachada": largura_fachada2,
            "uso": uso2,
            "zona_historica": zona_hist2_val,
            "n_pisos_favt1_for_rule": n_pisos1,
        }, tables, ref, "FAVT2")

        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            st.markdown("### Resultado FAVT 1")
            st.json(res1)
            pdf1 = render_pdf(res1, "Relatório de Simulação - FAVT 1 (Original)")
            st.download_button("Download PDF (FAVT1)", data=pdf1, file_name="favt1_resultado.pdf", mime="application/pdf")
        with c2:
            st.markdown("### Resultado FAVT 2")
            st.json(res2)
            pdf2 = render_pdf(res2, "Relatório de Simulação - FAVT 2 (Alterado)")
            st.download_button("Download PDF (FAVT2)", data=pdf2, file_name="favt2_resultado.pdf", mime="application/pdf")
        with c3:
            st.markdown("### Comparativo")
            delta = res2["valor_terreno"] - res1["valor_terreno"]
            st.metric("Δ Valor do Terreno ($)", f"{delta:,.2f}")
            st.metric("Valor FAVT1 ($)", f"{res1['valor_terreno']:,.2f}")
            st.metric("Valor FAVT2 ($)", f"{res2['valor_terreno']:,.2f}")

    except Exception as e:
        st.error(f"Erro no cálculo: {e}")
        st.info("Se o erro referir 'Zona/Bairro', confirma se a coluna 'Zona' no Excel corresponde ao agrupamento territorial.")
